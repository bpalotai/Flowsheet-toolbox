import os
import json
import pandas as pd
import numpy as np
from flask import Blueprint, request, jsonify, send_file, current_app, render_template, flash, redirect, url_for
from datetime import datetime
import tempfile
import matplotlib
matplotlib.use('Agg')  # Use the 'Agg' backend which doesn't require a GUI
import matplotlib.pyplot as plt
import io
import pythoncom
# Import visualization libraries
import seaborn as sns
import base64

from api.database import db, Case, SimulationResult
from driver.Sampling.sampling_driver import SamplingDriver

# Create blueprint
sampling_bp = Blueprint('sampling', __name__, url_prefix='/sampling')

@sampling_bp.route('/')
def sampling_index():
    # Fetch all sampling sets from the database
    samplings = SimulationResult.query.filter(
        SimulationResult.simulation_type.like('Sampling%')
    ).all()
    return render_template('sampling/sampling_landing.html', samplings=samplings, active_page='sampling')

@sampling_bp.route('/create', methods=['GET'])
def create_sampling():
    """Render the sampling page"""
    cases = Case.query.all()
    return render_template('sampling/sampling.html', cases=cases, active_page='sampling')

@sampling_bp.route('/parameters/<int:case_id>', methods=['GET'])
def get_case_parameters(case_id):
    """Get available parameters for a case"""
    case = Case.query.get_or_404(case_id)
    
    # Format the parameters for the frontend
    result = {
        'inputs': {},
        'outputs': {}
    }
    
    if case.parameters:
        # Process input parameters
        if 'InputParams' in case.parameters:
            result['inputs'] = case.parameters['InputParams']
        
        # Process output parameters
        if 'OutputParameters' in case.parameters:
            result['outputs'] = case.parameters['OutputParameters']
    
    return jsonify(result)


@sampling_bp.route('/generate', methods=['POST'])
def generate_samples():
    """Generate samples based on input parameters"""
    data = request.json
    
    # Extract parameters from request
    case_id = data.get('case_id')
    method = data.get('method', 'lhs')
    num_samples = data.get('num_samples', 100)
    input_params = data.get('input_params', [])
    param_ranges = data.get('param_ranges', {})
    
    # Check if the method is supported
    supported_methods = SamplingDriver.DRIVERS.keys()
    if method not in supported_methods:
        return jsonify({
            "error": f"Sampling method '{method}' is not currently supported. Supported methods are: {', '.join(supported_methods)}"
        }), 400

    if not case_id or not input_params or not param_ranges:
        return jsonify({"error": "Missing required parameters"}), 400
    
    # Get case
    case = Case.query.get_or_404(case_id)
    
    try:
        # Prepare parameter ranges for sampling driver
        sampling_ranges = {}
        # For Monte Carlo, we'll also collect distribution information
        distributions = {}

        for param_name in input_params:
            if param_name in param_ranges:
                param_config = param_ranges[param_name]
                
                if param_config['type'] == 'scalar':
                    # Regular parameter with min/max
                    sampling_ranges[param_name] = {
                        'min': param_config['min'],
                        'max': param_config['max']
                    }

                    # Store distribution type for Monte Carlo
                    if method == 'montecarlo' and 'distribution' in param_config:
                        distributions[param_name] = param_config['distribution']

                elif param_config['type'] == 'component':
                    # Component-based parameter
                    # For component parameters, we'll handle them separately
                    # We'll create a separate parameter for each component
                    for component, comp_config in param_config['components'].items():
                        component_param = f"{param_name}_{component}"
                        sampling_ranges[component_param] = {
                            'min': comp_config['min'],
                            'max': comp_config['max']
                        }
                        # Store distribution type for Monte Carlo
                        if method == 'montecarlo' and 'distribution' in comp_config:
                            distributions[component_param] = comp_config['distribution']
        
        # Create sampling driver with additional parameters for Monte Carlo
        if method == 'montecarlo':
            sampler = SamplingDriver(method, sampling_ranges, random_seed=42, distributions=distributions)
        else:
            sampler = SamplingDriver(method, sampling_ranges, random_seed=42)
        
        # Generate samples
        samples_df = sampler.generate_samples(num_samples)
        
        # Process component parameters
        for param_name in input_params:
            if param_name in param_ranges and param_ranges[param_name]['type'] == 'component':
                # Get all components for this parameter
                components = param_ranges[param_name]['components']
                
                # For each sample, create the component dictionary
                for i in range(len(samples_df)):
                    component_dict = {}
                    
                    for component in components:
                        component_param = f"{param_name}_{component}"
                        if component_param in samples_df.columns:
                            component_dict[component] = samples_df.at[i, component_param]
                            
                    # If we're dealing with mass fractions, normalize them
                    if param_ranges[param_name].get('is_fraction', False) or 'MassFrac' in param_name:
                        total = sum(component_dict.values())
                        if total > 0:
                            for component in component_dict:
                                component_dict[component] /= total
                    
                    # Store the component dictionary
                    samples_df.at[i, param_name] = component_dict
                
                # Remove individual component columns
                for component in components:
                    component_param = f"{param_name}_{component}"
                    if component_param in samples_df.columns:
                        samples_df = samples_df.drop(columns=[component_param])
        
        # Convert DataFrame to list of dictionaries
        samples_list = samples_df.to_dict(orient='records')
        
        # Return the samples
        return jsonify({
            "samples": samples_list,
            "method": method,
            "num_samples": num_samples
        })
    except Exception as e:
        current_app.logger.error(f"Error generating samples: {str(e)}")
        return jsonify({"error": f"Error generating samples: {str(e)}"}), 500

@sampling_bp.route('/save', methods=['POST'])
def save_sampling():
    """Save sampling results to the database"""
    data = request.json
    
    # Extract parameters from request
    case_id = data.get('case_id')
    name = data.get('name')
    description = data.get('description', '')
    method = data.get('method', 'lhs')
    num_samples = data.get('num_samples', 100)
    input_params = data.get('input_params', [])
    output_params = data.get('output_params', [])
    param_ranges = data.get('param_ranges', {})
    samples = data.get('samples', [])
    
    if not case_id or not name or not samples:
        return jsonify({"error": "Missing required parameters"}), 400
    
    # Get case
    case = Case.query.get_or_404(case_id)
    
    try:
        # Create a timestamp for the file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Create directory for sampling results if it doesn't exist
        sampling_dir = os.path.join('Cases', case.name, 'SamplingResults')
        os.makedirs(sampling_dir, exist_ok=True)
        
        # Create Excel file path
        excel_path = os.path.join(sampling_dir, f"sampling_{timestamp}.xlsx")
        
        # Create a DataFrame from the samples
        samples_df = pd.DataFrame(samples)
        
        # Create Excel writer
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Write samples to Excel
            samples_df.to_excel(writer, sheet_name='Samples', index=True)
            
            # Write metadata to Excel
            metadata = {
                'Name': [name],
                'Description': [description],
                'Method': [method],
                'Number of Samples': [num_samples],
                'Timestamp': [timestamp],
                'Case ID': [case_id],
                'Case Name': [case.name]
            }
            
            pd.DataFrame(metadata).to_excel(writer, sheet_name='Metadata', index=False)
            
            # Write parameter ranges to Excel
            ranges_data = []
            
            for param_name, param_config in param_ranges.items():
                if param_config['type'] == 'scalar':
                    ranges_data.append({
                        'Parameter': param_name,
                        'Type': 'Scalar',
                        'Min': param_config['min'],
                        'Max': param_config['max'],
                        'Distribution': param_config.get('distribution', 'uniform')
                    })
                elif param_config['type'] == 'component':
                    for component, comp_config in param_config['components'].items():
                        ranges_data.append({
                            'Parameter': f"{param_name} - {component}",
                            'Type': 'Component',
                            'Min': comp_config['min'],
                            'Max': comp_config['max'],
                            'Distribution': comp_config.get('distribution', 'uniform')
                        })
            
            pd.DataFrame(ranges_data).to_excel(writer, sheet_name='ParameterRanges', index=False)
        
        # Create a new simulation result entry
        result = SimulationResult(
            case_id=case_id,
            name=name,
            simulation_type=f"Sampling ({method})",
            data={
                'method': method,
                'num_samples': num_samples,
                'input_params': input_params,
                'output_params': output_params,
                'param_ranges': param_ranges,
                'samples_path': excel_path,
                'timestamp': timestamp
            }
        )
        
        # Add to database
        db.session.add(result)
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "Sampling saved successfully",
            "sampling_id": result.id,
            "file_path": excel_path
        })
    except Exception as e:
        current_app.logger.error(f"Error saving sampling: {str(e)}")
        return jsonify({"error": f"Error saving sampling: {str(e)}"}), 500

@sampling_bp.route('/download', methods=['POST'])
def download_sampling():
    """Download sampling results as Excel file"""
    data = request.json
    
    # Extract parameters from request
    case_id = data.get('case_id')
    case_name = data.get('case_name', 'Case')
    method = data.get('method', 'lhs')
    num_samples = data.get('num_samples', 100)
    samples = data.get('samples', [])
    
    if not samples:
        return jsonify({"error": "No samples to download"}), 400
    
    try:
        # Create a temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name
        
        # Create a DataFrame from the samples
        samples_df = pd.DataFrame(samples)
        
        # Create Excel writer
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            # Write samples to Excel
            samples_df.to_excel(writer, sheet_name='Samples', index=True)
            
            # Write metadata to Excel
            metadata = {
                'Method': [method],
                'Number of Samples': [num_samples],
                'Timestamp': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                'Case ID': [case_id],
                'Case Name': [case_name]
            }
            
            pd.DataFrame(metadata).to_excel(writer, sheet_name='Metadata', index=False)
            
            # Create a visualization sheet
            if len(samples_df.columns) >= 2:
                # Create scatter plot of first two parameters
                plt.figure(figsize=(10, 6))
                plt.scatter(samples_df.iloc[:, 0], samples_df.iloc[:, 1], alpha=0.7)
                plt.xlabel(samples_df.columns[0])
                plt.ylabel(samples_df.columns[1])
                plt.title(f"{method.upper()} Sampling")
                plt.grid(True, linestyle='--', alpha=0.7)
                
                # Save plot to buffer
                buf = io.BytesIO()
                plt.savefig(buf, format='png')
                plt.close()
                
                # Add image to Excel
                from openpyxl import load_workbook
                from openpyxl.drawing.image import Image
                
                # Save the Excel file first
                writer.save()
                
                # Then load it and add the image
                wb = load_workbook(temp_path)
                ws = wb.create_sheet('Visualization')
                
                # Reset buffer position
                buf.seek(0)
                
                # Add image to worksheet
                img = Image(buf)
                img.width = 600
                img.height = 400
                ws.add_image(img, 'B2')
                
                # Save workbook
                wb.save(temp_path)
        
        # Send the file
        return send_file(
            temp_path,
            as_attachment=True,
            download_name=f"Sampling_{case_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        current_app.logger.error(f"Error downloading sampling: {str(e)}")
        return jsonify({"error": f"Error downloading sampling: {str(e)}"}), 500
    finally:
        # Clean up temporary file
        if os.path.exists(temp_path):
            os.unlink(temp_path)

@sampling_bp.route('/run_simulations', methods=['POST'])
def run_simulations():
    """Run simulations with the generated samples"""
    data = request.json
    
    # Extract parameters from request
    case_id = data.get('case_id')
    samples = data.get('samples', [])
    flatten_components = data.get('flatten_components', False)  # New option with default True
    
    
    if not case_id or not samples:
        return jsonify({"error": "Missing required parameters"}), 400
    
    # Get case
    case = Case.query.get_or_404(case_id)
    
    try:
        # Create a timestamp for the batch
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Create directory for batch results if it doesn't exist
        batch_dir = os.path.join('Cases', case.name, 'BatchResults')
        os.makedirs(batch_dir, exist_ok=True)
        
        # Create a batch ID
        batch_id = f"batch_{timestamp}"
        
        # Create batch directory
        batch_path = os.path.join(batch_dir, batch_id)
        os.makedirs(batch_path, exist_ok=True)
        
        # Run simulations for each sample
        results = []
        completed = 0
        
        # Store static inputs - these are inputs that don't change between samples
        static_inputs = {}
        has_static_inputs = False
        
        for i, sample in enumerate(samples):
            try:
                # Use the existing simulation API to run each simulation
                from flask import current_app
                with current_app.test_client() as client:
                    response = client.post(
                        f'/simulations/api/run/{case_id}',
                        json={'input_params': sample}
                    )
                    
                    if response.status_code == 200:
                        sim_result = response.get_json()
                        # Get all inputs from simulation result
                        all_inputs = sim_result.get('results', {}).get('inputs', {})
                        
                        # Identify static inputs (inputs that are not in the sample)
                        if i == 0:  # Only need to do this for the first successful simulation
                            static_inputs = {k: v for k, v in all_inputs.items() if k not in sample}
                            if static_inputs:
                                has_static_inputs = True
                        
                        # Store result
                        results.append({
                            'sample_id': i,
                            'inputs': sample,
                            'static_inputs': {k: v for k, v in all_inputs.items() if k not in sample},
                            'outputs': sim_result.get('results', {}).get('outputs', {}),
                            'success': True
                        })
                        
                        completed += 1
                    else:
                        # Simulation failed
                        error_data = response.get_json()
                        results.append({
                            'sample_id': i,
                            'inputs': sample, 
                            'static_inputs': {},
                            'outputs': {},
                            'error': error_data.get('error', 'Unknown error'),
                            'success': False
                        })
            except Exception as e:
                current_app.logger.error(f"Error running simulation for sample {i}: {str(e)}")
                results.append({
                    'sample_id': i,
                    'inputs': sample,
                    'static_inputs': {},
                    'outputs': {},
                    'error': str(e),
                    'success': False
                })
        
        # Save results to Excel
        results_path = os.path.join(batch_path, 'batch_results.xlsx')
        
        # Process static inputs if we have any
        static_input_data = []
        if has_static_inputs and static_inputs:
            for key, value in static_inputs.items():
                # Check if this is a component parameter (a dictionary without value/uom at the top level)
                if isinstance(value, dict) and not ('value' in value and 'uom' in value):
                    # This is a component dictionary
                    # Extract component values into a simplified dictionary
                    simplified_components = {}
                    component_unit = ""
                    
                    for comp_name, comp_value in value.items():
                        if isinstance(comp_value, dict) and 'value' in comp_value:
                            simplified_components[comp_name] = comp_value['value']
                            # Use the same unit for all components (typically they all have the same unit)
                            if not component_unit and 'uom' in comp_value:
                                component_unit = comp_value['uom']
                        else:
                            simplified_components[comp_name] = comp_value
                    
                    # Add as a single row with the simplified component dictionary
                    static_input_data.append({
                        'Parameter': key,
                        'Value': str(simplified_components),  # Convert to string for Excel
                        'Unit': component_unit
                    })
                else:
                    # Regular parameter
                    param_val = value
                    param_unit = ""
                    
                    if isinstance(value, dict) and 'value' in value:
                        param_val = value['value']
                        param_unit = value.get('uom', '')
                    
                    static_input_data.append({
                        'Parameter': key,
                        'Value': param_val,
                        'Unit': param_unit
                    })


        # Create DataFrames for inputs and outputs
        input_data = []
        output_data = []

        for result in results:
            if result['success']:
                # Add input parameters
                input_row = {'Sample ID': result['sample_id']}
                
                # Process input parameters
                for key, value in result['inputs'].items():
                    # Check if this is a component parameter
                    if isinstance(value, dict) and not ('value' in value and 'uom' in value):
                        # For component dictionaries, either flatten or store as JSON
                        if flatten_components:
                            # Flatten: create separate columns for each component
                            for comp_name, comp_value in value.items():
                                comp_val = comp_value
                                if isinstance(comp_value, dict) and 'value' in comp_value:
                                    comp_val = comp_value['value']
                                input_row[f"{key}_{comp_name}"] = comp_val
                        else:
                            # Store as formatted string for better readability
                            formatted_components = {}
                            for comp_name, comp_value in value.items():
                                if isinstance(comp_value, dict) and 'value' in comp_value:
                                    formatted_components[comp_name] = comp_value['value']
                                else:
                                    formatted_components[comp_name] = comp_value
                            input_row[key] = json.dumps(formatted_components)
                    else:
                        # Regular parameter
                        if isinstance(value, dict) and 'value' in value:
                            input_row[key] = value['value']
                        else:
                            input_row[key] = value
                
                input_data.append(input_row)
                
                # Add output parameters
                output_row = {'Sample ID': result['sample_id']}
                
                # Process output parameters
                for key, value in result['outputs'].items():
                    # Check if this is a component parameter
                    if isinstance(value, dict) and not ('value' in value and 'uom' in value):
                        # For component dictionaries, either flatten or store as JSON
                        if flatten_components:
                            # Flatten: create separate columns for each component
                            for comp_name, comp_value in value.items():
                                comp_val = comp_value
                                if isinstance(comp_value, dict) and 'value' in comp_value:
                                    comp_val = comp_value['value']
                                output_row[f"{key}_{comp_name}"] = comp_val
                        else:
                            # Store as formatted string for better readability
                            formatted_components = {}
                            for comp_name, comp_value in value.items():
                                if isinstance(comp_value, dict) and 'value' in comp_value:
                                    formatted_components[comp_name] = comp_value['value']
                                else:
                                    formatted_components[comp_name] = comp_value
                            output_row[key] = json.dumps(formatted_components)
                    else:
                        # Regular parameter
                        if isinstance(value, dict) and 'value' in value:
                            output_row[key] = value['value']
                        else:
                            output_row[key] = value
                
                output_data.append(output_row)
        
        # Get the sampling method from the request data
        sampling_method = request.json.get('method', 'unknown')

        # Create Excel writer
        with pd.ExcelWriter(results_path, engine='openpyxl') as writer:
            # Convert to DataFrames
            input_df = pd.DataFrame(input_data)
            output_df = pd.DataFrame(output_data)
            
            # Write to Excel
            input_df.to_excel(writer, sheet_name='Inputs', index=False)
            output_df.to_excel(writer, sheet_name='Outputs', index=False)
            
            # Write static inputs to Excel if available
            if has_static_inputs and static_input_data:
                static_input_df = pd.DataFrame(static_input_data)
                static_input_df.to_excel(writer, sheet_name='Static Inputs', index=False)

            # Write metadata to Excel
            metadata = {
                'Batch ID': [batch_id],
                'Case ID': [case_id],
                'Case Name': [case.name],
                'Total Samples': [len(samples)],
                'Sampling Method': [sampling_method],
                'Completed Simulations': [completed],
                'Timestamp': [timestamp],
                'Flatten Components': [flatten_components],
                'Has Static Inputs': [has_static_inputs]
            }
            
            pd.DataFrame(metadata).to_excel(writer, sheet_name='Metadata', index=False)

        # Create a new simulation result entry for the batch
        result = SimulationResult(
            case_id=case_id,
            name=f"Batch Simulation ({sampling_method}) - {completed}/{len(samples)} completed",
            simulation_type="Batch Run",
            data={
                'batch_id': batch_id,
                'total_samples': len(samples),
                'completed': completed,
                'results_path': results_path,
                'timestamp': timestamp,
                'sampling_method': sampling_method,  # Store the original sampling method
                'flatten_components': flatten_components,
                'has_static_inputs': has_static_inputs,
                'static_inputs': static_input_data if has_static_inputs else None,
                'raw_data': {
                    'inputs': input_data,
                    'outputs': output_data
                }
            }
        )
        
        # Add to database
        db.session.add(result)
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"Batch simulation completed: {completed}/{len(samples)} simulations successful",
            "batch_id": batch_id,
            "result_id": result.id,
            "completed": completed,
            "total": len(samples)
        })
    except Exception as e:
        current_app.logger.error(f"Error running simulations: {str(e)}")
        return jsonify({"error": f"Error running simulations: {str(e)}"}), 500

@sampling_bp.route('/components/<int:case_id>', methods=['GET'])
def get_components(case_id):
    """Get components for a material stream parameter"""
    case = Case.query.get_or_404(case_id)
    
    # Get stream name and parameter type from query parameters
    stream_name = request.args.get('stream')
    param_type = request.args.get('param_type')
    
    if not stream_name or not param_type:
        return jsonify({"error": "Stream name and parameter type are required"}), 400
    
    try:
        # Check if we have component data in the case parameters
        if case.parameters:
            # Look for component data in input parameters
            if 'InputParams' in case.parameters and 'MaterialStream' in case.parameters['InputParams']:
                for param_name, param_config in case.parameters['InputParams']['MaterialStream'].items():
                    if param_config.get('StreamName') == stream_name and param_config.get('ParameterType') == param_type:
                        # If we have component data for this parameter, return it
                        if 'Components' in param_config:
                            components = param_config['Components']
                            
                            # Determine unit based on parameter type
                            unit = param_config.get('UOM', '')
                            if not unit:
                                if param_type == 'MassFrac':
                                    unit = 'fraction'
                                elif param_type == 'MassFlow':
                                    unit = 'kg/h'
                                elif param_type == 'MolarFlow':
                                    unit = 'kgmole/h'
                            
                            return jsonify({
                                "components": components,
                                "unit": unit
                            })
            
            # If not found in input parameters, check output parameters
            if 'OutputParameters' in case.parameters and 'MaterialStream' in case.parameters['OutputParameters']:
                for param_name, param_config in case.parameters['OutputParameters']['MaterialStream'].items():
                    if param_config.get('StreamName') == stream_name and param_config.get('ParameterType') == param_type:
                        # If we have component data for this parameter, return it
                        if 'Components' in param_config:
                            components = param_config['Components']
                            
                            # Determine unit based on parameter type
                            unit = param_config.get('UOM', '')
                            if not unit:
                                if param_type == 'MassFrac':
                                    unit = 'fraction'
                                elif param_type == 'MassFlow':
                                    unit = 'kg/h'
                                elif param_type == 'MolarFlow':
                                    unit = 'kgmole/h'
                            
                            return jsonify({
                                "components": components,
                                "unit": unit
                            })
        
        # If we don't have component data in the case parameters, return an empty object
        return jsonify({
            "components": {},
            "unit": ""
        })
            
    except Exception as e:
        current_app.logger.error(f"Error getting components: {str(e)}")
        return jsonify({"error": f"Error getting components: {str(e)}"}), 500


@sampling_bp.route('/results/<int:case_id>', methods=['GET'])
def get_sampling_results(case_id):
    """Get all sampling results for a case"""
    case = Case.query.get_or_404(case_id)
    
    # Get all simulation results of type 'Sampling'
    results = SimulationResult.query.filter_by(
        case_id=case_id,
        simulation_type='Sampling (lhs)'
    ).all()
    
    # Add other sampling types
    for method in ['Sampling (grid)', 'Sampling (montecarlo)', 'Sampling (sobol)']:
        results.extend(SimulationResult.query.filter_by(
            case_id=case_id,
            simulation_type=method
        ).all())
    
    # Format results
    formatted_results = []
    
    for result in results:
        formatted_results.append({
            'id': result.id,
            'name': result.name,
            'timestamp': result.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'type': result.simulation_type,
            'data': result.data
        })
    
    return jsonify(formatted_results)

@sampling_bp.route('/result/<int:result_id>', methods=['GET'])
def get_sampling_result(result_id):
    """Get details of a specific sampling result"""
    result = SimulationResult.query.get_or_404(result_id)
    
    # Check if this is a sampling result
    if 'Sampling' not in result.simulation_type:
        return jsonify({"error": "Not a sampling result"}), 400
    
    # Get the samples from the Excel file
    samples_path = result.data.get('samples_path')
    
    if not samples_path or not os.path.exists(samples_path):
        return jsonify({"error": "Samples file not found"}), 404
    
    try:
        # Read samples from Excel
        samples_df = pd.read_excel(samples_path, sheet_name='Samples', index_col=0)
        
        # Convert to list of dictionaries
        samples = samples_df.to_dict(orient='records')
        
        # Format result
        formatted_result = {
            'id': result.id,
            'name': result.name,
            'timestamp': result.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'type': result.simulation_type,
            'data': result.data,
            'samples': samples
        }
        
        return jsonify(formatted_result)
    except Exception as e:
        current_app.logger.error(f"Error reading samples: {str(e)}")
        return jsonify({"error": f"Error reading samples: {str(e)}"}), 500

@sampling_bp.route('/delete/<int:result_id>', methods=['DELETE'])
def delete_sampling_result(result_id):
    """Delete a sampling result"""
    result = SimulationResult.query.get_or_404(result_id)
    
    # Check if this is a sampling result
    if 'Sampling' not in result.simulation_type:
        return jsonify({"error": "Not a sampling result"}), 400
    
    try:
        # Delete the Excel file
        samples_path = result.data.get('samples_path')
        if samples_path and os.path.exists(samples_path):
            os.remove(samples_path)
        
        # Delete from database
        db.session.delete(result)
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "Sampling result deleted successfully"
        })
    except Exception as e:
        current_app.logger.error(f"Error deleting sampling result: {str(e)}"), 500
        return jsonify({"error": f"Error deleting sampling result: {str(e)}"}), 500

@sampling_bp.route('/batch/<int:result_id>', methods=['GET'])
def get_batch_result(result_id):
    """Get details of a batch simulation result"""
    result = SimulationResult.query.get_or_404(result_id)
    
    # Check if this is a batch result
    if result.simulation_type != "Batch Run":
        return jsonify({"error": "Not a batch simulation result"}), 400
    
    # Get the results from the Excel file
    results_path = result.data.get('results_path')
    
    if not results_path or not os.path.exists(results_path):
        return jsonify({"error": "Results file not found"}), 404
    
    try:
        # Read inputs and outputs from Excel
        inputs_df = pd.read_excel(results_path, sheet_name='Inputs', index_col='Sample ID')
        outputs_df = pd.read_excel(results_path, sheet_name='Outputs', index_col='Sample ID')
        
        # Read static inputs if available
        has_static_inputs = result.data.get('has_static_inputs', False)
        static_inputs = None
        
        if has_static_inputs:
            try:
                static_inputs_df = pd.read_excel(results_path, sheet_name='Static Inputs', index_col=None)
                static_inputs = static_inputs_df.to_dict(orient='records')
            except Exception as e:
                current_app.logger.warning(f"Could not read static inputs: {str(e)}")
                has_static_inputs = False

        # Convert to list of dictionaries
        inputs = inputs_df.to_dict(orient='records')
        outputs = outputs_df.to_dict(orient='records')
        
        # Format result
        formatted_result = {
            'id': result.id,
            'name': result.name,
            'timestamp': result.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'type': result.simulation_type,
            'data': result.data,
            'inputs': inputs,
            'outputs': outputs,
            'static_inputs': static_inputs,
            'has_static_inputs': has_static_inputs
        }
        
        return jsonify(formatted_result)
    except Exception as e:
        current_app.logger.error(f"Error reading batch results: {str(e)}")
        return jsonify({"error": f"Error reading batch results: {str(e)}"}), 500

# Helper function to extract numeric values from dictionary-like strings
def extract_numeric_value(value):
    """Extract numeric value from a dictionary-like string or object"""
    if isinstance(value, (int, float)):
        return value
    
    if isinstance(value, str):
        # Check if it's a dictionary-like string
        if value.startswith('{') and value.endswith('}'):
            try:
                # Try to parse as a dictionary
                import ast
                parsed = ast.literal_eval(value)
                
                # If it's a dictionary with a 'value' key, return that value
                if isinstance(parsed, dict) and 'value' in parsed:
                    return float(parsed['value'])
            except:
                pass
    
    # If all else fails, try to convert directly to float
    try:
        return float(value)
    except:
        return value

# Helper function to check if a column is numeric
def is_numeric_column(column):
    """Check if a pandas Series contains numeric values"""
    try:
        # Try to convert to numeric and check if any non-NaN values exist
        numeric_column = pd.to_numeric(column, errors='coerce')
        return not numeric_column.dropna().empty
    except:
        return False

@sampling_bp.route('/methods', methods=['GET'])
def get_sampling_methods():
    """Get available sampling methods"""
    methods = [
        {
            'id': 'lhs',
            'name': 'Latin Hypercube Sampling (LHS)',
            'description': 'Stratified sampling method that ensures good coverage of the parameter space.'
        },
        {
            'id': 'grid',
            'name': 'Grid Sampling',
            'description': 'Systematic sampling method that creates a regular grid across the parameter space.'
        },
        {
            'id': 'montecarlo',
            'name': 'Monte Carlo',
            'description': 'Random sampling method that generates independent samples from the parameter space.'
        },
        {
            'id': 'sobol',
            'name': 'Sobol Sequence',
            'description': 'Quasi-random low-discrepancy sequence that provides good coverage of the parameter space.'
        }
    ]
    
    return jsonify(methods)

@sampling_bp.route('/details/<int:result_id>')
def sampling_details(result_id):
    """Render the sampling details page"""
    # Get the sampling result
    result = SimulationResult.query.get_or_404(result_id)
    
    # Check if this is a sampling result
    if 'Sampling' not in result.simulation_type:
        flash('Not a sampling result', 'danger')
        return redirect(url_for('sampling.index'))
    
    # Get the case
    case = Case.query.get_or_404(result.case_id)
    
    # Get the samples from the Excel file
    samples_path = result.data.get('samples_path')
    
    if not samples_path or not os.path.exists(samples_path):
        flash('Samples file not found', 'danger')
        return redirect(url_for('sampling.index'))
    
    try:
        # Read samples from Excel (just metadata, not all samples)
        metadata_df = pd.read_excel(samples_path, sheet_name='Metadata')
        metadata = metadata_df.to_dict(orient='records')[0] if not metadata_df.empty else {}
        
        # Read parameter ranges
        try:
            ranges_df = pd.read_excel(samples_path, sheet_name='ParameterRanges')
            param_ranges = ranges_df.to_dict(orient='records')
        except:
            param_ranges = []
        
        # Get sample count without loading all samples
        sample_count = 0
        try:
            with pd.ExcelFile(samples_path) as xls:
                sample_count = len(pd.read_excel(xls, sheet_name='Samples', nrows=0).columns)
        except:
            pass
        
        return render_template(
            'sampling/details.html',
            result=result,
            case=case,
            metadata=metadata,
            param_ranges=param_ranges,
            sample_count=sample_count,
            active_page='sampling',
            referrer=request.referrer,
        )
    except Exception as e:
        current_app.logger.error(f"Error reading sampling details: {str(e)}")
        flash(f"Error reading sampling details: {str(e)}", 'danger')
        return redirect(url_for('sampling.index'))

@sampling_bp.route('/download/<int:result_id>')
def download_sampling_result(result_id):
    """Download a sampling result as Excel file"""
    result = SimulationResult.query.get_or_404(result_id)
    
    # Check if this is a sampling result
    if 'Sampling' not in result.simulation_type:
        flash('Not a sampling result', 'danger')
        return redirect(url_for('sampling.index'))
    
    # Get the samples from the Excel file
    samples_path = result.data.get('samples_path')
    
    if not samples_path or not os.path.exists(samples_path):
        flash('Samples file not found', 'danger')
        return redirect(url_for('sampling.index'))
    
    try:
        # Generate filename
        filename = f"Sampling_{result.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Send the file
        return send_file(
            samples_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        current_app.logger.error(f"Error downloading sampling result: {str(e)}")
        flash(f"Error downloading sampling result: {str(e)}", 'danger')
        return redirect(url_for('sampling.index'))

@sampling_bp.route('/batch/<int:result_id>/download')
def download_batch_result(result_id):
    """Download a batch simulation result as Excel file"""
    result = SimulationResult.query.get_or_404(result_id)
    
    # Check if this is a batch result
    if result.simulation_type != "Batch Run":
        flash('Not a batch simulation result', 'danger')
        return redirect(url_for('cases.case_detail', case_id=result.case_id))
    
    # Get the results from the Excel file
    results_path = result.data.get('results_path')
    
    if not results_path or not os.path.exists(results_path):
        flash('Results file not found', 'danger')
        return redirect(url_for('cases.case_detail', case_id=result.case_id))
    
    try:
        # Generate filename
        filename = f"Batch_Simulation_{result.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Send the file
        return send_file(
            results_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        current_app.logger.error(f"Error downloading batch results: {str(e)}")
        flash(f"Error downloading batch results: {str(e)}", 'danger')
        return redirect(url_for('cases.case_detail', case_id=result.case_id))

@sampling_bp.route('/batch/visualize/<int:result_id>', methods=['GET'])
def batch_visualize_page(result_id):
    """Render the batch visualization page"""
    result = SimulationResult.query.get_or_404(result_id)
    case = Case.query.get_or_404(result.case_id)
    
    # Check if this is a batch result
    if result.simulation_type != "Batch Run":
        flash("Not a batch simulation result", "danger")
        return redirect(url_for('simulations.simulation_results', case_id=case.id))
    
    return render_template('simulations/batch_visualize.html', 
                          result=result, 
                          case=case, 
                          active_page='cases')

@sampling_bp.route('/batch/<int:result_id>/parameters', methods=['GET'])
def get_batch_parameters(result_id):
    """Get available parameters for a batch simulation result"""
    result = SimulationResult.query.get_or_404(result_id)
    
    # Check if this is a batch result
    if result.simulation_type != "Batch Run":
        return jsonify({"error": "Not a batch simulation result"}), 400
    
    # Get the results from the Excel file
    results_path = result.data.get('results_path')
    
    if not results_path or not os.path.exists(results_path):
        return jsonify({"error": "Results file not found"}), 404
    
    try:
        # Read inputs and outputs from Excel
        inputs_df = pd.read_excel(results_path, sheet_name='Inputs', index_col=None)
        outputs_df = pd.read_excel(results_path, sheet_name='Outputs', index_col=None)
        
        # Get parameter names
        input_params = [col for col in inputs_df.columns if col != 'Sample ID']
        output_params = [col for col in outputs_df.columns if col != 'Sample ID']
        
        return jsonify({
            "input_params": input_params,
            "output_params": output_params
        })
    except Exception as e:
        current_app.logger.error(f"Error reading batch parameters: {str(e)}")
        return jsonify({"error": f"Failed to read batch parameters: {str(e)}"}), 500

@sampling_bp.route('/batch/<int:result_id>/visualize', methods=['GET'])
def visualize_batch_result(result_id):
    """Generate visualizations for a batch simulation result"""
    result = SimulationResult.query.get_or_404(result_id)
    
    # Check if this is a batch result
    if result.simulation_type != "Batch Run":
        return jsonify({"error": "Not a batch simulation result"}), 400
    
    # Get the results from the Excel file
    results_path = result.data.get('results_path')
    
    if not results_path or not os.path.exists(results_path):
        return jsonify({"error": "Results file not found"}), 404
    
    try:
        # Read inputs and outputs from Excel
        inputs_df = pd.read_excel(results_path, sheet_name='Inputs', index_col=None)
        outputs_df = pd.read_excel(results_path, sheet_name='Outputs', index_col=None)
        
        # Get visualization type from request
        viz_type = request.args.get('type', 'heatmap')
        
        # Get parameters from request
        input_params = request.args.getlist('input_params')
        output_params = request.args.getlist('output_params')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 6))  # Number of visualizations per page
        
        # Merge inputs and outputs on Sample ID
        if 'Sample ID' in inputs_df.columns and 'Sample ID' in outputs_df.columns:
            merged_df = pd.merge(inputs_df, outputs_df, on='Sample ID', suffixes=('_input', '_output'))
        else:
            # If no Sample ID column, just use the first columns
            merged_df = pd.concat([inputs_df, outputs_df], axis=1)
        
        # Get all available parameters for selection UI
        all_input_params = [col for col in inputs_df.columns if col != 'Sample ID']
        all_output_params = [col for col in outputs_df.columns if col != 'Sample ID']
        
        # Initialize response data with empty defaults for all visualization types
        response_data = {
            'success': True,
            'parameters': {
                'input_params': all_input_params,
                'output_params': all_output_params,
                'selected_inputs': input_params,
                'selected_outputs': output_params
            },
            'visualizations': [],
            'correlation_heatmap': None,
            'contribution_analysis': [],
            'statistics': {},
            'pagination': {
                'current_page': 1,
                'total_pages': 1,
                'total_plots': 0,
                'per_page': per_page
            }
        }
        
        # Set seaborn style
        sns.set(style="whitegrid")
        
        # Generate visualizations based on type
        if viz_type == 'heatmap':
            # Generate correlation heatmap
            response_data['correlation_heatmap'] = generate_correlation_heatmap(merged_df)
            
        elif viz_type == 'scatter':
            # Generate scatter plots
            scatter_data = generate_scatter_plots(merged_df, input_params, output_params)
            response_data['visualizations'] = scatter_data.get('visualizations', [])
            response_data['pagination'] = scatter_data.get('pagination', response_data['pagination'])
            
        elif viz_type == 'contribution':
            # Generate contribution analysis
            response_data['contribution_analysis'] = generate_contribution_analysis(merged_df, input_params, output_params)
            
        elif viz_type == 'statistics':
            # Generate statistics
            response_data['statistics'] = generate_statistics(merged_df)
        
        else:
            # If no specific type is requested, generate all visualizations
            response_data['correlation_heatmap'] = generate_correlation_heatmap(merged_df)
            
            scatter_data = generate_scatter_plots(merged_df, input_params, output_params, page, per_page)
            response_data['visualizations'] = scatter_data.get('visualizations', [])
            response_data['pagination'] = scatter_data.get('pagination', response_data['pagination'])
            
            response_data['contribution_analysis'] = generate_contribution_analysis(merged_df, input_params, output_params)
            response_data['statistics'] = generate_statistics(merged_df)
        
        return jsonify(response_data)
    except Exception as e:
        current_app.logger.error(f"Error generating visualizations: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "error": f"Error generating visualizations: {str(e)}",
            "visualizations": [],
            "correlation_heatmap": None,
            "contribution_analysis": [],
            "statistics": {},
            "pagination": {
                "current_page": 1,
                "total_pages": 1,
                "total_plots": 0,
                "per_page": per_page
            }
        }), 500

def generate_correlation_heatmap(df):
    """Generate correlation heatmap for all numeric parameters"""
    # Filter to only include numeric columns for correlation
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    
    if len(numeric_cols) <= 1:
        return None
    
    plt.figure(figsize=(12, 10))
    
    # Calculate correlation matrix
    corr_matrix = df[numeric_cols].corr()
    
    # Create heatmap
    mask = np.triu(np.ones_like(corr_matrix, dtype=bool))
    sns.heatmap(
        corr_matrix, 
        mask=mask,
        cmap='coolwarm', 
        annot=True, 
        fmt=".2f",
        linewidths=0.5,
        cbar_kws={"shrink": .8}
    )
    
    plt.title('Parameter Correlation Heatmap', fontsize=16)
    plt.tight_layout()
    
    # Save plot to buffer
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100)
    plt.close()
    
    # Convert to base64 for embedding in HTML
    buf.seek(0)
    img_str = base64.b64encode(buf.read()).decode('utf-8')
    
    return {
        'image': img_str,
        'description': 'Correlation matrix between all numeric parameters. Values range from -1 (perfect negative correlation) to 1 (perfect positive correlation).'
    }

def generate_scatter_plots(df, input_params, output_params, page=1, per_page=6):
    """Generate scatter plots for selected input and output parameters"""
    visualizations = []
    
    # If specific parameters are provided, use only those
    if input_params and output_params:
        input_cols = [col for col in input_params if col in df.columns]
        output_cols = [col for col in output_params if col in df.columns]
    else:
        # Otherwise, use all numeric columns
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        
        # Try to identify input and output columns based on naming conventions
        input_cols = [col for col in numeric_cols if col.startswith('Input') or col in df.columns[:len(df.columns)//2]]
        output_cols = [col for col in numeric_cols if col.startswith('Output') or col in df.columns[len(df.columns)//2:]]
    
    # Create scatter plots for each input-output pair
    plot_combinations = []
    for input_col in input_cols:
        if input_col == 'Sample ID':
            continue
            
        for output_col in output_cols:
            if output_col == 'Sample ID':
                continue
                
            # Skip if either column has non-numeric values
            if not is_numeric_column(df[input_col]) or not is_numeric_column(df[output_col]):
                continue
            
            plot_combinations.append((input_col, output_col))
    
    # Calculate total pages
    total_plots = len(plot_combinations)
    total_pages = (total_plots + per_page - 1) // per_page if total_plots > 0 else 1
    
    # Apply pagination if multiple plots
    if len(plot_combinations) > 0:
        start_idx = (page - 1) * per_page
        end_idx = min(start_idx + per_page, total_plots)
        current_combinations = plot_combinations[start_idx:end_idx]
    else:
        current_combinations = []
    
    # Generate plots for current page
    for input_col, output_col in current_combinations:
        try:
            # Create scatter plot with regression line using seaborn
            plt.figure(figsize=(10, 6))
            
            # Create seaborn scatter plot with regression line
            ax = sns.regplot(
                x=input_col, 
                y=output_col, 
                data=df,
                scatter_kws={"alpha": 0.6, "s": 50},
                line_kws={"color": "red"}
            )
            
            # Calculate correlation coefficient
            corr = df[[input_col, output_col]].corr().iloc[0, 1]
            
            # Add correlation coefficient to the plot
            plt.annotate(
                f"Correlation: {corr:.4f}",
                xy=(0.05, 0.95),
                xycoords='axes fraction',
                fontsize=12,
                bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.8)
            )
            
            plt.title(f"{output_col} vs {input_col}", fontsize=14)
            plt.xlabel(input_col, fontsize=12)
            plt.ylabel(output_col, fontsize=12)
            plt.tight_layout()
            
            # Save plot to buffer
            buf = io.BytesIO()
            plt.savefig(buf, format='png', dpi=100)
            plt.close()
            
            # Convert to base64 for embedding in HTML
            buf.seek(0)
            img_str = base64.b64encode(buf.read()).decode('utf-8')
            
            visualizations.append({
                'title': f"{output_col} vs {input_col}",
                'type': 'scatter',
                'image': img_str,
                'correlation': float(corr) if not np.isnan(corr) else None
            })
        except Exception as e:
            current_app.logger.error(f"Error generating scatter plot for {input_col} vs {output_col}: {str(e)}")
            # Add a placeholder for the failed visualization
            visualizations.append({
                'title': f"{output_col} vs {input_col}",
                'type': 'scatter',
                'error': str(e),
                'correlation': None
            })
    
    # If no visualizations were generated but parameters were provided, add an error message
    if len(visualizations) == 0 and (input_params or output_params):
        return {
            'visualizations': [{
                'title': 'Error',
                'type': 'error',
                'error': 'Could not generate visualizations for the selected parameters. They may not be numeric or contain insufficient data.'
            }],
            'pagination': {
                'current_page': page,
                'total_pages': total_pages,
                'total_plots': total_plots,
                'per_page': per_page
            }
        }
    
    return {
        'visualizations': visualizations,
        'pagination': {
            'current_page': page,
            'total_pages': total_pages,
            'total_plots': total_plots,
            'per_page': per_page
        }
    }


def generate_contribution_analysis(df, input_params=None, output_params=None):
    """Generate contribution analysis for selected output parameters"""
    contribution_analysis = []
    
    # Determine which columns to use
    if input_params:
        input_cols = [col for col in input_params if col in df.columns]
    else:
        # Try to identify input columns based on naming conventions or position
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        input_cols = [col for col in numeric_cols if col.startswith('Input') or col in df.columns[:len(df.columns)//2]]
    
    if output_params:
        output_cols = [col for col in output_params if col in df.columns]
    else:
        # Try to identify output columns based on naming conventions or position
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        output_cols = [col for col in numeric_cols if col.startswith('Output') or col in df.columns[len(df.columns)//2:]]
    
    # For each output parameter, calculate correlation with all input parameters
    for output_col in output_cols:
        if output_col == 'Sample ID':
            continue
            
        # Skip if output column has non-numeric values
        if not is_numeric_column(df[output_col]):
            continue
            
        # Calculate correlation with each input parameter
        correlations = []
        for input_col in input_cols:
            if input_col == 'Sample ID':
                continue
                
            # Skip if input column has non-numeric values
            if not is_numeric_column(df[input_col]):
                continue
            
            # Calculate correlation
            corr = df[[input_col, output_col]].corr().iloc[0, 1]
            correlations.append((input_col, float(corr) if not np.isnan(corr) else None))
        
        # Sort by absolute correlation value
        correlations.sort(key=lambda x: abs(x[1]) if x[1] is not None else 0, reverse=True)
        
        # Create contribution plot (top 10 inputs)
        if correlations:
            plt.figure(figsize=(12, 6))
            
            # Extract top correlations (up to 10)
            top_correlations = correlations[:10]
            
            # Create DataFrame for plotting
            corr_df = pd.DataFrame(top_correlations, columns=['Parameter', 'Correlation'])
            
            # Filter out None values
            corr_df = corr_df.dropna()
            
            if len(corr_df) > 0:
                # Create horizontal bar plot
                ax = sns.barplot(
                    x='Correlation',
                    y='Parameter',
                    data=corr_df,
                    palette='viridis'
                )
                
                # Add value labels
                for i, v in enumerate(corr_df['Correlation']):
                    ax.text(v + (0.01 if v >= 0 else -0.01), 
                            i, 
                            f"{v:.4f}", 
                            va='center',
                            fontweight='bold',
                            color='black' if v >= 0 else 'white')
                
                plt.title(f"Input Parameters Contribution to {output_col}", fontsize=14)
                plt.xlabel('Correlation Coefficient', fontsize=12)
                plt.ylabel('Input Parameter', fontsize=12)
                plt.axvline(x=0, color='gray', linestyle='--')
                plt.tight_layout()
                
                # Save plot to buffer
                buf = io.BytesIO()
                plt.savefig(buf, format='png', dpi=100)
                plt.close()
                
                # Convert to base64 for embedding in HTML
                buf.seek(0)
                img_str = base64.b64encode(buf.read()).decode('utf-8')
                
                contribution_analysis.append({
                    'output': output_col,
                    'image': img_str,
                    'correlations': correlations
                })
    
    return contribution_analysis

def generate_statistics(df):
    """Generate statistics for all numeric parameters"""
    statistics = {}
    
    # Get numeric columns
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    
    for col in numeric_cols:
        if col == 'Sample ID':
            continue
            
        values = df[col].dropna()
        
        if len(values) > 0:
            statistics[col] = {
                'min': float(values.min()),
                'max': float(values.max()),
                'mean': float(values.mean()),
                'median': float(values.median()),
                'std': float(values.std())
            }
            
            # Handle NaN values
            for key, value in statistics[col].items():
                if np.isnan(value):
                    statistics[col][key] = None
    
    return statistics
